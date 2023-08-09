import os
import openpyxl

from tkinter import Tk, filedialog

# Función para obtener la lista de elementos en una carpeta
def obtener_lista_elementos(carpeta):
    elementos = os.listdir(carpeta)
    return elementos

# Función para crear un archivo Excel con la lista de elementos
def crear_excel(lista_elementos):
    libro_excel = openpyxl.Workbook()
    hoja = libro_excel.active
    hoja.title = "Lista de Elementos"
    
    hoja.cell(row=1, column=1, value="Nombre de Elemento")
    
    for idx, elemento in enumerate(lista_elementos, start=2):
        hoja.cell(row=idx, column=1, value=elemento)
    
    return libro_excel

# Diálogo para seleccionar una carpeta
root = Tk()
root.withdraw()  # Ocultar la ventana principal

carpeta_seleccionada = filedialog.askdirectory(title="Selecciona una carpeta")

if carpeta_seleccionada:
    lista_elementos = obtener_lista_elementos(carpeta_seleccionada)
    
    libro_excel = crear_excel(lista_elementos)
    excel_file_path = os.path.join(carpeta_seleccionada, "lista_elementos.xlsx")
    libro_excel.save(excel_file_path)
    
    print(f"Archivo Excel creado en: {excel_file_path}")
else:
    print("No se seleccionó ninguna carpeta.")
