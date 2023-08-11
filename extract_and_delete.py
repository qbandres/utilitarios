import os
import openpyxl
from tkinter import Tk, filedialog, Label, Listbox, Button, Scrollbar, StringVar, IntVar
from tkinter import messagebox

# Función para obtener la lista de elementos en una carpeta
def obtener_lista_elementos(carpeta):
    elementos = os.listdir(carpeta)
    return elementos

# Función para crear un archivo Excel con la lista de elementos
def crear_excel(lista_elementos):
    libro_excel = openpyxl.Workbook()
    hoja = libro_excel.active
    hoja.title = "Lista de Elementos"
    
    hoja.cell(row=1, column=1, value="Nombre de Elemento")
    
    for idx, elemento in enumerate(lista_elementos, start=2):
        hoja.cell(row=idx, column=1, value=elemento)
    
    return libro_excel

# Función para borrar elementos seleccionados
def borrar_elementos_seleccionados():
    elementos_seleccionados = elementos_listbox.curselection()
    
    if elementos_seleccionados:
        elementos_a_borrar = [elementos_listbox.get(idx) for idx in elementos_seleccionados]
        confirmar_borrado = messagebox.askyesno("Confirmar borrado", f"¿Seguro que quieres borrar estos elementos?\n{elementos_a_borrar}")
        
        if confirmar_borrado:
            for elemento in elementos_a_borrar:
                ruta_elemento = os.path.join(carpeta_seleccionada, elemento)
                if os.path.isfile(ruta_elemento):
                    os.remove(ruta_elemento)
                elif os.path.isdir(ruta_elemento):
                    os.rmdir(ruta_elemento)
            actualizar_lista_elementos()

def actualizar_lista_elementos():
    lista_elementos = obtener_lista_elementos(carpeta_seleccionada)
    elementos_listbox.delete(0, 'end')
    
    for elemento in lista_elementos:
        elementos_listbox.insert('end', elemento)

def mostrar_elementos_por_tipo():
    tipos = {}
    for elemento in obtener_lista_elementos(carpeta_seleccionada):
        tipo = "Archivo" if os.path.isfile(os.path.join(carpeta_seleccionada, elemento)) else "Carpeta"
        if tipo not in tipos:
            tipos[tipo] = []
        tipos[tipo].append(elemento)
    
    elementos_listbox.delete(0, 'end')
    for tipo, elementos in tipos.items():
        elementos_listbox.insert('end', tipo)
        for elemento in elementos:
            elementos_listbox.insert('end', f"  {elemento}")

# Crear la ventana principal de tkinter
root = Tk()
root.title("Gestor de Archivos")

carpeta_seleccionada = ""

# Label y Listbox para mostrar la lista de elementos
label = Label(root, text="Selecciona una carpeta:")
label.pack()

elementos_listbox = Listbox(root, selectmode="multiple", width=50, height=15)
elementos_listbox.pack()

# Scrollbar para el Listbox
scrollbar = Scrollbar(root)
scrollbar.pack(side="right", fill="y")
elementos_listbox.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=elementos_listbox.yview)

# Botones para acciones
boton_seleccionar = Button(root, text="Seleccionar Carpeta", command=lambda: seleccionar_carpeta())
boton_seleccionar.pack()

boton_mostrar_por_tipo = Button(root, text="Mostrar por Tipo", command=lambda: mostrar_elementos_por_tipo())
boton_mostrar_por_tipo.pack()

boton_borrar = Button(root, text="Borrar Seleccionados", command=lambda: borrar_elementos_seleccionados())
boton_borrar.pack()

# Función para seleccionar una carpeta
def seleccionar_carpeta():
    global carpeta_seleccionada
    carpeta_seleccionada = filedialog.askdirectory(title="Selecciona una carpeta")
    
    if carpeta_seleccionada:
        actualizar_lista_elementos()
    else:
        print("No se seleccionó ninguna carpeta.")

# Iniciar el bucle de tkinter
root.mainloop()
